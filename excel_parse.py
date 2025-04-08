import re
from openpyxl import load_workbook

def preprocess_text(text):
    """
    Làm sạch văn bản: loại bỏ dấu nháy đơn, dấu câu và khoảng trắng thừa.
    """
    # Loại bỏ dấu nháy đơn
    text = text.replace("'", "")
    
    # Loại bỏ dấu câu như :, !, ?
    text = re.sub(r'[^\w\s]', '', text)
    
    # Loại bỏ khoảng trắng thừa
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text

def Extract_excel_data(file_path: str):
    extracted_data = []
    extracted_value = []
    workbook = load_workbook(file_path)
    sheet = workbook.active  # Lấy sheet đầu tiên trong workbook
    
    # Lấy danh sách các vùng merged cell
    merged_ranges = list(sheet.merged_cells.ranges)
    
    # Hàm kiểm tra xem cell có nằm trong vùng merged cell hay không
    def get_merged_range(cell, merged_ranges):
        for merged_range in merged_ranges:
            if cell.coordinate in merged_range:
                return merged_range
        return None
    
    # Duyệt qua từng ô trong sheet
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = repr(cell.value)
            # Làm sạch văn bản trước khi lưu
            clean_value = preprocess_text(value)
            # Loại bỏ giá trị nếu chỉ toàn các chữ "n" (một hoặc nhiều chữ)
            if re.fullmatch(r"n+", clean_value):
                continue
            extracted_value.append(clean_value)
    
            coord = cell.coordinate
            
            # Kiểm tra xem cell có nằm trong vùng merged cell hay không
            merged_range = get_merged_range(cell, merged_ranges)
            if merged_range:
                merged_info = f"{merged_range}"
                extracted_data.append(merged_info)
            else:
                # Nếu không nằm trong merged cell, ta trả về tọa độ cell
                extracted_data.append(coord)
    
    # Tạo một dict nối extracted_data và extracted_value theo cặp key-value
    combined_dict = dict(zip(extracted_data, extracted_value))
    
    return extracted_data, extracted_value, combined_dict

# Kiểm tra kết quả
data, values, combined = Extract_excel_data('invoice_empty.xlsx')



